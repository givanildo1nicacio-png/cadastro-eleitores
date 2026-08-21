#!/usr/bin/env python3
"""
Cadastro de Eleitores - Sistema de cadastro online
Funciona em computador e celular (design responsivo)
Com autenticação de usuário (login/senha)
"""

import os
import sqlite3
import secrets
from datetime import datetime
from functools import wraps
from flask import (Flask, render_template, request, jsonify,
                   redirect, url_for, session, flash)

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', secrets.token_hex(32))

# Database: usa /tmp no Render (único diretório gravável)
_db_dir = os.environ.get('RENDER_DISK', os.path.dirname(os.path.abspath(__file__)))
DATABASE = os.path.join(_db_dir, 'eleitores.db')


# ─── Banco de Dados ───────────────────────────────────────────

def get_db():
    """Conecta ao banco SQLite"""
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    """Cria as tabelas se não existirem"""
    conn = get_db()

    # Tabela de usuários
    conn.execute('''
        CREATE TABLE IF NOT EXISTS usuarios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL UNIQUE,
            senha_hash TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # Tabela de eleitores
    conn.execute('''
        CREATE TABLE IF NOT EXISTS eleitores (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            titulo_eleitoral TEXT NOT NULL UNIQUE,
            numero_titulo TEXT,
            nome_completo TEXT NOT NULL,
            data_nascimento TEXT,
            sexo TEXT,
            telefone TEXT,
            cep TEXT,
            logradouro TEXT,
            numero TEXT,
            complemento TEXT,
            bairro TEXT,
            cidade TEXT,
            estado TEXT,
            zona_eleitoral TEXT,
            secao_eleitoral TEXT,
            observacoes TEXT,
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (created_by) REFERENCES usuarios(id)
        )
    ''')
    conn.commit()
    conn.close()

    # Migração: remover coluna cpf se existir
    try:
        conn2 = get_db()
        conn2.execute('ALTER TABLE eleitores DROP COLUMN cpf')
        conn2.commit()
        conn2.close()
    except Exception:
        pass  # Coluna já removida ou não existe


# ─── Autenticação ─────────────────────────────────────────────

def hash_senha(senha):
    """Gera hash da senha com salt"""
    import hashlib
    salt = secrets.token_hex(16)
    h = hashlib.pbkdf2_hmac('sha256', senha.encode(), salt.encode(), 100000)
    return f"{salt}:{h.hex()}"


def verificar_senha(senha, hash_completo):
    """Verifica se a senha confere com o hash"""
    import hashlib
    salt, h = hash_completo.split(':')
    h_check = hashlib.pbkdf2_hmac('sha256', senha.encode(), salt.encode(), 100000)
    return h_check.hex() == h


def login_required(f):
    """Decorator que exige login"""
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'usuario_id' not in session:
            if request.path.startswith('/api/'):
                return jsonify({'erro': 'Acesso negado. Faça login.'}), 401
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


def get_usuario_logado():
    """Retorna dados do usuário logado"""
    if 'usuario_id' not in session:
        return None
    conn = get_db()
    user = conn.execute(
        'SELECT id, nome, email FROM usuarios WHERE id = ?',
        (session['usuario_id'],)
    ).fetchone()
    conn.close()
    return dict(user) if user else None


# ─── Rotas de Auth ────────────────────────────────────────────

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Página de login"""
    if request.method == 'POST':
        nome = request.form.get('nome', '').strip()
        senha = request.form.get('senha', '')

        if not nome or not senha:
            flash('Preencha nome e senha', 'error')
            return render_template('login.html')

        conn = get_db()
        user = conn.execute(
            'SELECT * FROM usuarios WHERE nome = ?', (nome,)
        ).fetchone()
        conn.close()

        if user and verificar_senha(senha, user['senha_hash']):
            session['usuario_id'] = user['id']
            session['usuario_nome'] = user['nome']
            return redirect(url_for('index'))
        else:
            flash('Nome ou senha incorretos', 'error')
            return render_template('login.html')

    return render_template('login.html')


@app.route('/registro', methods=['GET', 'POST'])
def registro():
    """Página de registro"""
    if request.method == 'POST':
        nome = request.form.get('nome', '').strip()
        senha = request.form.get('senha', '')
        confirmar = request.form.get('confirmar_senha', '')

        if not nome or not senha:
            flash('Preencha todos os campos', 'error')
            return render_template('registro.html')

        if len(senha) < 6:
            flash('A senha deve ter no mínimo 6 caracteres', 'error')
            return render_template('registro.html')

        if senha != confirmar:
            flash('As senhas não conferem', 'error')
            return render_template('registro.html')

        conn = get_db()
        try:
            conn.execute(
                'INSERT INTO usuarios (nome, senha_hash) VALUES (?, ?)',
                (nome, hash_senha(senha))
            )
            conn.commit()

            # Login automático após registro
            user = conn.execute(
                'SELECT * FROM usuarios WHERE nome = ?', (nome,)
            ).fetchone()
            conn.close()

            session['usuario_id'] = user['id']
            session['usuario_nome'] = user['nome']

            flash('Conta criada com sucesso!', 'success')
            return redirect(url_for('index'))
        except sqlite3.IntegrityError:
            conn.close()
            flash('Este nome de usuário já está em uso', 'error')
            return render_template('registro.html')

    return render_template('registro.html')


@app.route('/logout')
def logout():
    """Encerrar sessão"""
    session.clear()
    flash('Sessão encerrada', 'info')
    return redirect(url_for('login'))


# ─── Perfil do Usuário ────────────────────────────────────────

@app.route('/perfil', methods=['GET', 'POST'])
@login_required
def perfil():
    """Página de perfil do usuário"""
    conn = get_db()
    usuario = conn.execute(
        'SELECT id, nome, created_at FROM usuarios WHERE id = ?',
        (session['usuario_id'],)
    ).fetchone()

    # Estatísticas do usuário
    stats = conn.execute(
        'SELECT COUNT(*) as total FROM eleitores WHERE created_by = ?',
        (session['usuario_id'],)
    ).fetchone()
    conn.close()

    if request.method == 'POST':
        acao = request.form.get('acao', '')

        if acao == 'atualizar_perfil':
            nome = request.form.get('nome', '').strip()

            if not nome:
                flash('Nome é obrigatório', 'error')
                return render_template('perfil.html', usuario=dict(usuario), stats=dict(stats))

            conn = get_db()
            try:
                conn.execute(
                    'UPDATE usuarios SET nome = ? WHERE id = ?',
                    (nome, session['usuario_id'])
                )
                conn.commit()
                session['usuario_nome'] = nome
                flash('Perfil atualizado com sucesso!', 'success')
            except sqlite3.IntegrityError:
                flash('Este nome de usuário já está em uso', 'error')
            finally:
                conn.close()

            # Recarregar dados
            conn = get_db()
            usuario = conn.execute(
                'SELECT id, nome, created_at FROM usuarios WHERE id = ?',
                (session['usuario_id'],)
            ).fetchone()
            conn.close()

        elif acao == 'alterar_senha':
            senha_atual = request.form.get('senha_atual', '')
            nova_senha = request.form.get('nova_senha', '')
            confirmar = request.form.get('confirmar_senha', '')

            if not senha_atual or not nova_senha:
                flash('Preencha todos os campos de senha', 'error')
                return render_template('perfil.html', usuario=dict(usuario), stats=dict(stats))

            if len(nova_senha) < 6:
                flash('A nova senha deve ter no mínimo 6 caracteres', 'error')
                return render_template('perfil.html', usuario=dict(usuario), stats=dict(stats))

            if nova_senha != confirmar:
                flash('As senhas não conferem', 'error')
                return render_template('perfil.html', usuario=dict(usuario), stats=dict(stats))

            conn = get_db()
            user = conn.execute(
                'SELECT senha_hash FROM usuarios WHERE id = ?',
                (session['usuario_id'],)
            ).fetchone()
            conn.close()

            if not verificar_senha(senha_atual, user['senha_hash']):
                flash('Senha atual incorreta', 'error')
                return render_template('perfil.html', usuario=dict(usuario), stats=dict(stats))

            conn = get_db()
            conn.execute(
                'UPDATE usuarios SET senha_hash = ? WHERE id = ?',
                (hash_senha(nova_senha), session['usuario_id'])
            )
            conn.commit()
            conn.close()

            flash('Senha alterada com sucesso!', 'success')

    return render_template('perfil.html', usuario=dict(usuario), stats=dict(stats))


@app.route('/api/usuario/perfil', methods=['GET'])
@login_required
def api_perfil():
    """API: dados do perfil"""
    conn = get_db()
    user = conn.execute(
        'SELECT id, nome, created_at FROM usuarios WHERE id = ?',
        (session['usuario_id'],)
    ).fetchone()
    stats = conn.execute(
        'SELECT COUNT(*) as total FROM eleitores WHERE created_by = ?',
        (session['usuario_id'],)
    ).fetchone()
    conn.close()
    return jsonify({
        'id': user['id'],
        'nome': user['nome'],
        'created_at': user['created_at'],
        'eleitores_cadastrados': stats['total']
    })


# ─── Rotas Principais ─────────────────────────────────────────

@app.route('/')
@login_required
def index():
    """Página principal"""
    conn = get_db()
    eleitores = conn.execute(
        '''SELECT e.*, u.nome as criado_por_nome
           FROM eleitores e
           LEFT JOIN usuarios u ON e.created_by = u.id
           WHERE e.created_by = ?
           ORDER BY e.nome_completo ASC''',
        (session['usuario_id'],)
    ).fetchall()
    total = len(eleitores)
    conn.close()
    return render_template('index.html',
                           eleitores=eleitores,
                           total=total,
                           usuario=get_usuario_logado())


# ─── API Eleitores (protegida) ────────────────────────────────

@app.route('/api/eleitores', methods=['GET'])
@login_required
def api_listar():
    """API: listar todos os eleitores"""
    busca = request.args.get('busca', '').strip()
    conn = get_db()

    usuario_id = session.get('usuario_id')
    if busca:
        eleitores = conn.execute(
            '''SELECT e.*, u.nome as criado_por_nome
               FROM eleitores e
               LEFT JOIN usuarios u ON e.created_by = u.id
               WHERE e.created_by = ?
                 AND (e.nome_completo LIKE ?
                      OR e.titulo_eleitoral LIKE ?
                      OR e.cidade LIKE ?
                      OR e.bairro LIKE ?)
               ORDER BY e.nome_completo ASC''',
            (usuario_id, f'%{busca}%', f'%{busca}%', f'%{busca}%', f'%{busca}%')
        ).fetchall()
    else:
        eleitores = conn.execute(
            '''SELECT e.*, u.nome as criado_por_nome
               FROM eleitores e
               LEFT JOIN usuarios u ON e.created_by = u.id
               WHERE e.created_by = ?
               ORDER BY e.nome_completo ASC''',
            (usuario_id,)
        ).fetchall()

    conn.close()
    return jsonify([dict(e) for e in eleitores])


@app.route('/api/eleitores', methods=['POST'])
@login_required
def api_cadastrar():
    """API: cadastrar novo eleitor"""
    data = request.get_json()

    if not data.get('titulo_eleitoral') or not data.get('nome_completo'):
        return jsonify({'erro': 'Nº do Título e nome são obrigatórios'}), 400

    conn = get_db()
    try:
        conn.execute('''
        INSERT INTO eleitores (
            titulo_eleitoral, numero_titulo, nome_completo, data_nascimento,
            sexo, telefone, cep, logradouro, numero,
            complemento, bairro, cidade, estado,
            zona_eleitoral, secao_eleitoral, observacoes, created_by
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            data.get('titulo_eleitoral', '').strip(),
            data.get('numero_titulo', '').strip(),
            data.get('nome_completo', '').strip(),
            data.get('data_nascimento', '').strip(),
            data.get('sexo', '').strip(),
            data.get('telefone', '').strip(),
            data.get('cep', '').strip(),
            data.get('logradouro', '').strip(),
            data.get('numero', '').strip(),
            data.get('complemento', '').strip(),
            data.get('bairro', '').strip(),
            data.get('cidade', '').strip(),
            data.get('estado', '').strip(),
            data.get('zona_eleitoral', '').strip(),
            data.get('secao_eleitoral', '').strip(),
            data.get('observacoes', '').strip(),
            session.get('usuario_id'),
        ))
        conn.commit()
        return jsonify({'sucesso': True, 'mensagem': 'Eleitor cadastrado com sucesso!'}), 201
    except sqlite3.IntegrityError:
        return jsonify({'erro': 'Este título eleitoral já está cadastrado'}), 409
    finally:
        conn.close()



@app.route('/api/eleitores/<int:id>', methods=['GET'])
@login_required
def api_consultar(id):
    """API: consultar um eleitor"""
    conn = get_db()
    eleitor = conn.execute(
        '''SELECT e.*, u.nome as criado_por_nome
           FROM eleitores e
           LEFT JOIN usuarios u ON e.created_by = u.id
           WHERE e.id = ?''', (id,)
    ).fetchone()
    conn.close()
    if eleitor:
        return jsonify(dict(eleitor))
    return jsonify({'erro': 'Eleitor não encontrado'}), 404


@app.route('/api/eleitores/<int:id>', methods=['PUT'])
@login_required
def api_atualizar(id):
    """API: atualizar eleitor"""
    data = request.get_json()
    conn = get_db()

    eleitor = conn.execute('SELECT * FROM eleitores WHERE id = ?', (id,)).fetchone()
    if not eleitor:
        conn.close()
        return jsonify({'erro': 'Eleitor não encontrado'}), 404

    conn.execute('''
        UPDATE eleitores SET
            titulo_eleitoral = ?, numero_titulo = ?, nome_completo = ?,
            data_nascimento = ?, sexo = ?, telefone = ?,
            cep = ?, logradouro = ?, numero = ?, complemento = ?,
            bairro = ?, cidade = ?, estado = ?,
            zona_eleitoral = ?, secao_eleitoral = ?, observacoes = ?,
            updated_at = ?
        WHERE id = ?
    ''', (
        data.get('titulo_eleitoral', '').strip(),
        data.get('numero_titulo', '').strip(),
        data.get('nome_completo', '').strip(),
        data.get('data_nascimento', '').strip(),
        data.get('sexo', '').strip(),
        data.get('telefone', '').strip(),
        data.get('cep', '').strip(),
        data.get('logradouro', '').strip(),
        data.get('numero', '').strip(),
        data.get('complemento', '').strip(),
        data.get('bairro', '').strip(),
        data.get('cidade', '').strip(),
        data.get('estado', '').strip(),
        data.get('zona_eleitoral', '').strip(),
        data.get('secao_eleitoral', '').strip(),
        data.get('observacoes', '').strip(),
        datetime.now().isoformat(),
        id
    ))
    conn.commit()
    conn.close()
    return jsonify({'sucesso': True, 'mensagem': 'Eleitor atualizado com sucesso!'})


@app.route('/api/eleitores/<int:id>', methods=['DELETE'])
@login_required
def api_excluir(id):
    """API: excluir eleitor"""
    conn = get_db()
    eleitor = conn.execute('SELECT * FROM eleitores WHERE id = ?', (id,)).fetchone()
    if not eleitor:
        conn.close()
        return jsonify({'erro': 'Eleitor não encontrado'}), 404

    conn.execute('DELETE FROM eleitores WHERE id = ?', (id,))
    conn.commit()
    conn.close()
    return jsonify({'sucesso': True, 'mensagem': 'Eleitor excluído com sucesso!'})


@app.route('/api/stats', methods=['GET'])
@login_required
def api_stats():
    """API: estatísticas"""
    usuario_id = session.get('usuario_id')
    conn = get_db()
    total = conn.execute('SELECT COUNT(*) as c FROM eleitores WHERE created_by = ?', (usuario_id,)).fetchone()['c']
    por_estado = conn.execute(
        'SELECT estado, COUNT(*) as c FROM eleitores WHERE created_by = ? AND estado != "" GROUP BY estado ORDER BY c DESC',
        (usuario_id,)
    ).fetchall()
    conn.close()
    return jsonify({
        'total': total,
        'por_estado': [{'estado': e['estado'], 'count': e['c']} for e in por_estado],
        'por_usuario': [{'nome': u['nome'], 'count': u['c']} for u in por_usuario]
    })


# ─── Exportar PDF ────────────────────────────────────────────

@app.route('/api/export/pdf', methods=['GET'])
@login_required
def api_export_pdf():
    """Exporta lista de eleitores em PDF"""
    from fpdf import FPDF

    usuario_id = session.get('usuario_id')
    conn = get_db()
    eleitores = conn.execute(
        '''SELECT e.*, u.nome as criado_por_nome
           FROM eleitores e
           LEFT JOIN usuarios u ON e.created_by = u.id
           WHERE e.created_by = ?
           ORDER BY e.nome_completo ASC''',
        (usuario_id,)
    ).fetchall()
    conn.close()

    pdf = FPDF(orientation='L', unit='mm', format='A4')
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    # Título
    pdf.set_font('Helvetica', 'B', 16)
    pdf.cell(0, 10, 'Cadastro de Eleitores', ln=True, align='C')
    pdf.set_font('Helvetica', '', 10)
    pdf.cell(0, 8, f'Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}', ln=True, align='C')
    pdf.cell(0, 6, f'Total: {len(eleitores)} eleitores', ln=True, align='C')
    pdf.ln(8)

    # Cabeçalho da tabela
    colunas = [
        ('Nº Título', 28),
        ('Nome', 50),
        ('Telefone', 28),
        ('Cidade/UF', 35),
        ('Zona', 15),
        ('Seção', 15),
        ('Bairro', 35),
        ('Logradouro', 55),
    ]

    pdf.set_font('Helvetica', 'B', 7)
    pdf.set_fill_color(26, 82, 118)
    pdf.set_text_color(255, 255, 255)
    for titulo, largura in colunas:
        pdf.cell(largura, 8, titulo, border=1, fill=True, align='C')
    pdf.ln()

    # Dados
    pdf.set_font('Helvetica', '', 6.5)
    pdf.set_text_color(0, 0, 0)
    for i, e in enumerate(eleitores):
        if pdf.get_y() > 180:
            pdf.add_page()
            pdf.set_font('Helvetica', 'B', 7)
            pdf.set_fill_color(26, 82, 118)
            pdf.set_text_color(255, 255, 255)
            for titulo, largura in colunas:
                pdf.cell(largura, 8, titulo, border=1, fill=True, align='C')
            pdf.ln()
            pdf.set_font('Helvetica', '', 6.5)
            pdf.set_text_color(0, 0, 0)

        cor_fundo = (240, 245, 250) if i % 2 == 0 else (255, 255, 255)
        pdf.set_fill_color(*cor_fundo)

        cidade_uf = f"{e['cidade'] or ''}/{e['estado'] or ''}" if e['cidade'] else '-'
        zona_secao = f"{e['zona_eleitoral'] or ''}/{e['secao_eleitoral'] or ''}" if e['zona_eleitoral'] else '-'
        endereco = f"{e['logradouro'] or ''}, {e['numero'] or 'S/N'}" if e['logradouro'] else '-'

        dados = [
            e['numero_titulo'] or e['titulo_eleitoral'] or '-',
            (e['nome_completo'] or '-')[:35],
            e['telefone'] or '-',
            cidade_uf,
            e['zona_eleitoral'] or '-',
            e['secao_eleitoral'] or '-',
            e['bairro'] or '-',
            endereco[:40],
        ]

        for valor, (_, largura) in zip(dados, colunas):
            pdf.cell(largura, 6, str(valor)[:40], border=1, fill=True)
        pdf.ln()

    # Gerar response
    pdf_bytes = pdf.output()
    from io import BytesIO
    buf = BytesIO()
    buf.write(pdf_bytes)
    buf.seek(0)

    from flask import send_file
    return send_file(
        buf,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'eleitores_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf'
    )


# ─── Exportar Excel ──────────────────────────────────────────

@app.route('/api/export/excel', methods=['GET'])
@login_required
def api_export_excel():
    """Exporta lista de eleitores em Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO

    usuario_id = session.get('usuario_id')
    conn = get_db()
    eleitores = conn.execute(
        '''SELECT e.*, u.nome as criado_por_nome
           FROM eleitores e
           LEFT JOIN usuarios u ON e.created_by = u.id
           WHERE e.created_by = ?
           ORDER BY e.nome_completo ASC''',
        (usuario_id,)
    ).fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Eleitores'

    # Estilos
    header_font = Font(bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill(start_color='1A5276', end_color='1A5276', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Cabeçalhos
    cabecalhos = [
        'Nº Título', 'Nome Completo', 'Data Nasc.', 'Sexo',
        'Telefone', 'CEP', 'Logradouro', 'Número',
        'Complemento', 'Bairro', 'Cidade', 'Estado',
        'Zona Eleitoral', 'Seção Eleitoral', 'Observações',
        'Cadastrado por', 'Data Cadastro'
    ]

    for col, titulo in enumerate(cabecalhos, 1):
        cell = ws.cell(row=1, column=col, value=titulo)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    # Dados
    sexo_map = {'M': 'Masculino', 'F': 'Feminino', 'O': 'Outro'}
    for row_idx, e in enumerate(eleitores, 2):
        dados = [
            e['numero_titulo'] or e['titulo_eleitoral'],
            e['nome_completo'],
            e['data_nascimento'],
            sexo_map.get(e['sexo'], e['sexo']),
            e['telefone'],
            e['cep'],
            e['logradouro'],
            e['numero'],
            e['complemento'],
            e['bairro'],
            e['cidade'],
            e['estado'],
            e['zona_eleitoral'],
            e['secao_eleitoral'],
            e['observacoes'],
            e['criado_por_nome'],
            e['created_at'],
        ]
        for col, valor in enumerate(dados, 1):
            cell = ws.cell(row=row_idx, column=col, value=valor or '')
            cell.border = border
            cell.alignment = Alignment(vertical='center')

    # Ajustar largura das colunas
    larguras = [16, 30, 14, 12, 18, 12, 30, 8, 15, 20, 20, 8, 12, 12, 25, 18, 18]
    for i, largura in enumerate(larguras, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else chr(64 + (i-1)//26) + chr(65 + (i-1)%26)].width = largura

    # Congelar primeira linha
    ws.freeze_panes = 'A2'

    # Salvar em buffer
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    from flask import send_file
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'eleitores_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    )


# ─── Iniciar ──────────────────────────────────────────────────

# Sempre inicializa o banco ao importar (gunicorn importa o módulo)
init_db()

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    print("=" * 50)
    print("  CADASTRO DE ELEITORES")
    print(f"  Acesse: http://localhost:{port}")
    print("=" * 50)
    app.run(host='0.0.0.0', port=port, debug=True)
